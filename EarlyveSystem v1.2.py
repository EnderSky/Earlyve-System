# -*- coding: utf-8 -*- 

###########################################################################
## Python code generated with wxFormBuilder (version Jun 17 2015)
## http://www.wxformbuilder.org/
##
## PLEASE DO "NOT" EDIT THIS FILE!
###########################################################################

import wx
import wx.xrc
import datetime
import time
import openpyxl
import sys

###########################################################################
## Things Added from Build Code:
##   - Imported datetime, time, openpyxl and sys modules
##   - Added self.mode in __init__
##   - Functions: OnSubmit, OnHelp, OnAbout, OnUpdateLog, On Exit,
##     OnEarlyLeave, OnLatecomers
##   - Contents of About, Help and Update Log are read from external txt files
##   - Added self.Fit() to ABout, Help and Update Log Dialogs
##   - Run Code
###########################################################################

###########################################################################
## Class MainFrame
###########################################################################

class MainFrame ( wx.Frame ):
        
        def __init__( self, parent ):
                wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Earlyve System", pos = wx.DefaultPosition, size = wx.Size( 500,400 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
                
                self.SetSizeHintsSz( wx.DefaultSize, wx.DefaultSize )
                
                bSizerMainFrame = wx.BoxSizer( wx.VERTICAL )
                
                self.panelMain = wx.Panel( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.TAB_TRAVERSAL )
                bSizerMainFrame.Add( self.panelMain, 0, wx.EXPAND |wx.ALL, 0 )

                #Initialise mode as "Early Leave"
                self.mode = "EarlyLeave"
                
                self.textTitle = wx.StaticText( self, wx.ID_ANY, u"Early Leave", wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textTitle.Wrap( -1 )
                self.textTitle.SetFont( wx.Font( 15, 70, 90, 92, True, wx.EmptyString ) )
                
                bSizerMainFrame.Add( self.textTitle, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 15 )
                
                self.textEnterStudentID = wx.StaticText( self, wx.ID_ANY, u"Enter Student ID Below:", wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textEnterStudentID.Wrap( -1 )
                bSizerMainFrame.Add( self.textEnterStudentID, 0, wx.ALL, 10 )
                
                self.textfieldStudentID = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
                bSizerMainFrame.Add( self.textfieldStudentID, 0, wx.ALIGN_TOP|wx.ALL|wx.EXPAND, 10 )
                
                self.buttonSubmit = wx.Button( self, wx.ID_ANY, u"Submit", wx.DefaultPosition, wx.DefaultSize, 0 )
                bSizerMainFrame.Add( self.buttonSubmit, 0, wx.ALL|wx.EXPAND, 10 )
                
                self.textConfirmation = wx.StaticText( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textConfirmation.Wrap( -1 )
                bSizerMainFrame.Add( self.textConfirmation, 0, wx.ALIGN_CENTER|wx.ALIGN_TOP|wx.ALL|wx.EXPAND, 10 )
                
                
                self.SetSizer( bSizerMainFrame )
                self.Layout()
                self.menubarMain = wx.MenuBar( 0 )
                self.menuFile = wx.Menu()
                self.menuItemFileAbout = wx.MenuItem( self.menuFile, wx.ID_ANY, u"About", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuFile.AppendItem( self.menuItemFileAbout )
                
                self.menuItemFileHelp = wx.MenuItem( self.menuFile, wx.ID_ANY, u"Help", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuFile.AppendItem( self.menuItemFileHelp )
                
                self.menuItemFileUpdateLog = wx.MenuItem( self.menuFile, wx.ID_ANY, u"Update Log", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuFile.AppendItem( self.menuItemFileUpdateLog )
                
                self.menuFile.AppendSeparator()
                
                self.menuItemFileExit = wx.MenuItem( self.menuFile, wx.ID_ANY, u"Exit Program", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuFile.AppendItem( self.menuItemFileExit )
                
                self.menubarMain.Append( self.menuFile, u"File" ) 
                
                self.menuEarly = wx.Menu()
                self.menuItemEarlyEnterData = wx.MenuItem( self.menuEarly, wx.ID_ANY, u"Enter Student Data", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuEarly.AppendItem( self.menuItemEarlyEnterData )
                
                self.menubarMain.Append( self.menuEarly, u"Early Leave" ) 
                
                self.menuLate = wx.Menu()
                self.menuItemLateEnterData = wx.MenuItem( self.menuLate, wx.ID_ANY, u"Enter Student Data", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuLate.AppendItem( self.menuItemLateEnterData )
                
                self.menubarMain.Append( self.menuLate, u"Latecomers" ) 
                
                self.SetMenuBar( self.menubarMain )
                
                
                self.Centre( wx.BOTH )
                
                # Connect Events
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnSubmit )
                self.Bind( wx.EVT_MENU, self.OnAbout, id = self.menuItemFileAbout.GetId() )
                self.Bind( wx.EVT_MENU, self.OnHelp, id = self.menuItemFileHelp.GetId() )
                self.Bind( wx.EVT_MENU, self.OnUpdateLog, id = self.menuItemFileUpdateLog.GetId() )
                self.Bind( wx.EVT_MENU, self.OnExit, id = self.menuItemFileExit.GetId() )
                self.Bind( wx.EVT_MENU, self.OnEarlyLeave, id = self.menuItemEarlyEnterData.GetId() )
                self.Bind( wx.EVT_MENU, self.OnLatecomers, id = self.menuItemLateEnterData.GetId() )
        
        def __del__( self ):
                pass
        
        
        # Virtual event handlers, overide them in your derived class
        def OnSubmit( self, event ):
                #1. Get data (datetime and NRIC) and set confirmation text
                nric = self.textfieldStudentID.GetValue()
                curr_datetime = datetime.datetime.now()

                #Clear text field
                self.textfieldStudentID.Clear()
                
                self.textConfirmation.SetLabel("Student " + str(nric) + "'s data has been entered successfully into the system \nat " + str(curr_datetime))

                #time.sleep(0.5)
                #self.textConfirmation.SetLabel("")

                #2. Open Excel file and select correct sheet according to mode
                wb = openpyxl.load_workbook('Student Data.xlsx')
                
                if (self.mode == "EarlyLeave"):
                        sheet = wb["Early Leave"]
                elif (self.mode == "Latecomers"):
                        sheet = wb["Late"]

                #3. Check for empty row
                #Empty is the variable representing the first row empty
                empty = -1
                for i in range(1, sheet.max_row):

                        #If row is empty, set empty to that row number and break
                        if sheet['A'+str(i)].value is None:
                                empty = i
                                break
                    
                #if empty = -1, means that the rows are filled until sheet.max_row
                if empty == -1:
                        empty = sheet.max_row+1

                #4. Add data to Excel file and save
                #Set latest row of column A to datetime
                sheet['A' + str(empty)].value = curr_datetime
                #Set latest row of column B to student ID
                sheet['B' + str(empty)].value = nric
                wb.save('Student Data.xlsx')
        
        def OnAbout( self, event ):
                #Show About Dialog when About menu item selected
                AboutDialog(None).Show()
        
        def OnHelp( self, event ):
                #Show Help Dialog when Help menu item selected
                HelpDialog(None).Show()
        
        def OnUpdateLog( self, event ):
                #Show UpdateLog Dialog when Update Log menu item selected
                UpdateLogDialog(None).Show()
        
        def OnExit( self, event ):
                #Exit the program
                sys.exit()
        
        def OnEarlyLeave( self, event ):
                #Set mode as "EarlyLeave" and change title to "Early Leave"
                self.mode = "EarlyLeave"
                self.textTitle.SetLabel("Early Leave")
        
        def OnLatecomers( self, event ):
                #Set mode as "Latecomers" and change title to "Latecomers"
                self.mode = "Latecomers"
                self.textTitle.SetLabel("Latecomers")
        

###########################################################################
## Class AboutDialog
###########################################################################

class AboutDialog ( wx.Dialog ):
        
        def __init__( self, parent ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"About", pos = wx.DefaultPosition, size = wx.Size( 400,400 ), style = wx.DEFAULT_DIALOG_STYLE )
                
                self.SetSizeHintsSz( wx.DefaultSize, wx.DefaultSize )
                
                bSizerAbout = wx.BoxSizer( wx.VERTICAL )

                #Read About - Version.txt and add it to the dialog
                AboutVersion = open("About - Version.txt", "r")
                text = AboutVersion.read()
                AboutVersion.close()
                
                self.textAbout = wx.StaticText( self, wx.ID_ANY, text, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textAbout.Wrap( -1 )
                bSizerAbout.Add( self.textAbout, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 20 )

                #Read About - Features.txt and add it to the dialog
                AboutFeatures = open("About - Features.txt", "r")
                text = AboutFeatures.read()
                AboutFeatures.close()
                
                self.textFeatures = wx.StaticText( self, wx.ID_ANY, text, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textFeatures.Wrap( -1 )
                bSizerAbout.Add( self.textFeatures, 0, wx.ALL, 10 )
                
                
                self.SetSizer( bSizerAbout )
                self.Layout()
                
                #Added self.Fit()
                self.Fit()
                
                self.Centre( wx.BOTH )
        
        def __del__( self ):
                pass
        

###########################################################################
## Class HelpDialog
###########################################################################

class HelpDialog ( wx.Dialog ):
        
        def __init__( self, parent ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"Help", pos = wx.DefaultPosition, size = wx.Size( 500,300 ), style = wx.DEFAULT_DIALOG_STYLE )
                
                self.SetSizeHintsSz( wx.DefaultSize, wx.DefaultSize )
                
                bSizerHelp = wx.BoxSizer( wx.VERTICAL )
                
                self.textHelpTitle = wx.StaticText( self, wx.ID_ANY, u"Help", wx.DefaultPosition, wx.Size( -1,-1 ), 0 )
                self.textHelpTitle.Wrap( -1 )
                self.textHelpTitle.SetFont( wx.Font( 15, 70, 90, 92, False, wx.EmptyString ) )
                
                bSizerHelp.Add( self.textHelpTitle, 0, wx.ALIGN_LEFT|wx.ALL, 15 )

                #Read Help.txt and add it to the dialog
                HelpFile = open("Help.txt", "r")
                text = HelpFile.read()
                HelpFile.close()
                
                self.textHelpInstructions = wx.StaticText( self, wx.ID_ANY, text, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textHelpInstructions.Wrap( -1 )
                bSizerHelp.Add( self.textHelpInstructions, 0, wx.ALIGN_LEFT|wx.ALL, 10 )
                
                
                self.SetSizer( bSizerHelp )
                self.Layout()
                
                #Added self.Fit()
                self.Fit()
                
                self.Centre( wx.BOTH )
        
        def __del__( self ):
                pass
        

###########################################################################
## Class UpdateLogDialog
###########################################################################

class UpdateLogDialog ( wx.Dialog ):
        
        def __init__( self, parent ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"Help", pos = wx.DefaultPosition, size = wx.Size( 500,500 ), style = wx.DEFAULT_DIALOG_STYLE )
                
                self.SetSizeHintsSz( wx.DefaultSize, wx.DefaultSize )
                
                bSizerUpdateLog = wx.BoxSizer( wx.VERTICAL )
                
                self.textUpdateLogTitle = wx.StaticText( self, wx.ID_ANY, u"Update Log", wx.DefaultPosition, wx.Size( -1,-1 ), 0 )
                self.textUpdateLogTitle.Wrap( -1 )
                self.textUpdateLogTitle.SetFont( wx.Font( 15, 70, 90, 92, False, wx.EmptyString ) )
                
                bSizerUpdateLog.Add( self.textUpdateLogTitle, 0, wx.ALIGN_LEFT|wx.ALL, 15 )

                #Read Update Log.txt and add it to the dialog
                UpdateLog = open("Update Log.txt", "r")
                text = UpdateLog.read()
                UpdateLog.close()
                
                self.textUpdateLog = wx.StaticText( self, wx.ID_ANY, text, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textUpdateLog.Wrap( -1 )
                bSizerUpdateLog.Add( self.textUpdateLog, 0, wx.ALIGN_LEFT|wx.ALL, 10 )
                
                
                self.SetSizer( bSizerUpdateLog )
                self.Layout()
                
                #Added self.Fit()
                self.Fit()
                
                self.Centre( wx.BOTH )
        
        def __del__( self ):
                pass
        
#Run Code (allows app to be run)
class MainApp(wx.App):
        def OnInit(self):
                mainFrame = MainFrame(None)
                mainFrame.Show(True)
                return True

if __name__ == '__main__':
        app = MainApp()
        #Loop running of app, if not it will instantly disappear
        app.MainLoop()
