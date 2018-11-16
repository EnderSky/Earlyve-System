# -*- coding: utf-8 -*- 

###########################################################################
## Python code generated with wxFormBuilder (version Jun 17 2015)
## http://www.wxformbuilder.org/
##
## PLEASE DO "NOT" EDIT THIS FILE!
###########################################################################

import wx
import wx.lib.scrolledpanel as scrolled
import datetime
import time
import openpyxl
import sys
import csv
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

###########################################################################
## Global Variables
###########################################################################

file = open("Assets/Filename.txt", 'r')
filename = file.read()          #Filename of file which logs down the data
file.close()

#Last empty rows for Early Leave Excel sheet and Latecomers Excel Sheet
earlyleave_empty = 0
latecomers_empty = 0
###########################################################################
## Class MainFrame
###########################################################################

class ScrollPanel(scrolled.ScrolledPanel):

    def __init__(self, parent, title, text):

        scrolled.ScrolledPanel.__init__(self, parent, -1)

        ## Set up Panel and GUI
        bSizerPanel = wx.BoxSizer(wx.VERTICAL)
        
        textTitle = wx.StaticText( self, wx.ID_ANY, title, wx.DefaultPosition, wx.Size( -1,-1 ), 0 )
        textTitle.Wrap( -1 )
        textTitle.SetFont( wx.Font( 25, 70, 90, 92, False, wx.EmptyString ) )
        bSizerPanel.Add( textTitle, 0, wx.ALIGN_LEFT|wx.ALL, 15 )
        
        textInfo = wx.StaticText( self, wx.ID_ANY, text, wx.DefaultPosition, wx.DefaultSize, 0 )
        textInfo.Wrap( -1 )
        textInfo.SetFont( wx.Font( 12, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
        bSizerPanel.Add( textInfo, 0, wx.ALIGN_LEFT|wx.ALL|wx.EXPAND, 10 )

        #Format panel and set up scrolling
        self.SetSizer(bSizerPanel)
        self.SetupScrolling()
        self.Layout()


class MainFrame ( wx.Frame ):
        
        def __init__( self, parent ):
                wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Earlyve System", pos = wx.DefaultPosition, size = wx.Size( 500,400 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
                
                self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

                #Initialise mode as "Early Leave"
                self.mode = "EarlyLeave"

                #Initialise data
                self.FindEmpty()

                ## Set up Panel and GUI
                bSizerMainFrame = wx.BoxSizer( wx.VERTICAL )
                
                self.textTitle = wx.StaticText( self, wx.ID_ANY, u"Early Leave", wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textTitle.Wrap( -1 )
                self.textTitle.SetFont( wx.Font( 40, 70, 90, 92, True, wx.EmptyString ) )
                
                bSizerMainFrame.Add( self.textTitle, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 15 )
                
                self.textEnterStudentID = wx.StaticText( self, wx.ID_ANY, u"Enter Student ID Below:", wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textEnterStudentID.Wrap( -1 )
                self.textEnterStudentID.SetFont( wx.Font( 20, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerMainFrame.Add( self.textEnterStudentID, 0, wx.ALL, 10 )
                
                self.textfieldStudentID = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize )
                self.textfieldStudentID.SetFont( wx.Font( 25, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerMainFrame.Add( self.textfieldStudentID, 0, wx.ALL|wx.EXPAND, 10 )

                self.textReason = wx.StaticText( self, wx.ID_ANY, u"Reason for taking Early Leave:", wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textReason.Wrap( -1 )
                self.textReason.SetFont( wx.Font( 20, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerMainFrame.Add( self.textReason, 0, wx.ALL, 10 )

                self.textfieldReason = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize )
                self.textfieldReason.SetFont( wx.Font( 25, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerMainFrame.Add( self.textfieldReason, 0, wx.ALL|wx.EXPAND, 10 )
                
                self.buttonSubmit = wx.Button( self, wx.ID_ANY, u"Submit", wx.DefaultPosition, wx.DefaultSize, 0 )
                self.buttonSubmit.SetFont( wx.Font( 20, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerMainFrame.Add( self.buttonSubmit, 0, wx.ALL|wx.EXPAND, 10 )
                
                self.textConfirmation = wx.StaticText( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textConfirmation.SetFont( wx.Font( 15, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                self.textConfirmation.Wrap( -1 )
                bSizerMainFrame.Add( self.textConfirmation, 0, wx.ALL|wx.EXPAND, 15 )

                self.textEmailConfirmation = wx.StaticText( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textEmailConfirmation.SetFont( wx.Font( 15, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                self.textEmailConfirmation.Wrap( -1 )
                bSizerMainFrame.Add( self.textEmailConfirmation, 0, wx.ALL|wx.EXPAND, 15 )

                self.SetSize(1000,700)
                self.SetSizer( bSizerMainFrame )
                self.Layout()

                ## Set up Menubar
                self.menubarMain = wx.MenuBar( 0 )
                self.menuHome = wx.Menu()
                self.menuItemHomeAbout = wx.MenuItem( self.menuHome, wx.ID_ANY, u"About", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuHome.Append( self.menuItemHomeAbout )
                
                self.menuItemHomeHelp = wx.MenuItem( self.menuHome, wx.ID_ANY, u"Help", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuHome.Append( self.menuItemHomeHelp )

                self.menuItemHomeSetFilename = wx.MenuItem (self.menuHome, wx.ID_ANY, u"Set Filename", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuHome.Append( self.menuItemHomeSetFilename )
                
                self.menuItemHomeUpdateLog = wx.MenuItem( self.menuHome, wx.ID_ANY, u"Update Log", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuHome.Append( self.menuItemHomeUpdateLog )
                
                self.menuHome.AppendSeparator()
                
                self.menuItemHomeExit = wx.MenuItem( self.menuHome, wx.ID_ANY, u"Exit Program", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuHome.Append( self.menuItemHomeExit )
                
                self.menubarMain.Append( self.menuHome, u"Home" ) 
                
                self.menuEarly = wx.Menu()
                self.menuItemEarlyEnterData = wx.MenuItem( self.menuEarly, wx.ID_ANY, u"Enter Student Data", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuEarly.Append( self.menuItemEarlyEnterData )

                self.menuItemEarlyQueryID = wx.MenuItem( self.menuEarly, wx.ID_ANY, u"Query by Student ID", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuEarly.Append( self.menuItemEarlyQueryID )

                self.menuItemEarlyQueryDate = wx.MenuItem( self.menuEarly, wx.ID_ANY, u"Query by Date", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuEarly.Append( self.menuItemEarlyQueryDate )

                self.menuItemEarlyQueryName = wx.MenuItem( self.menuEarly, wx.ID_ANY, u"Query by Name", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuEarly.Append( self.menuItemEarlyQueryName )
                
                self.menubarMain.Append( self.menuEarly, u"Early Leave" ) 
                
                self.menuLate = wx.Menu()
                self.menuItemLateEnterData = wx.MenuItem( self.menuLate, wx.ID_ANY, u"Enter Student Data", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuLate.Append( self.menuItemLateEnterData )

                self.menuItemLateQueryID = wx.MenuItem( self.menuLate, wx.ID_ANY, u"Query by Student ID", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuLate.Append( self.menuItemLateQueryID )

                self.menuItemLateQueryDate = wx.MenuItem( self.menuLate, wx.ID_ANY, u"Query by Date", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuLate.Append( self.menuItemLateQueryDate )

                self.menuItemLateQueryName = wx.MenuItem( self.menuLate, wx.ID_ANY, u"Query by Name", wx.EmptyString, wx.ITEM_NORMAL )
                self.menuLate.Append( self.menuItemLateQueryName )
                
                self.menubarMain.Append( self.menuLate, u"Latecomers" ) 
                
                self.SetMenuBar( self.menubarMain )
                
                
                self.Centre( wx.BOTH )
                
                # Connect Events
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnSubmit )
                self.Bind( wx.EVT_MENU, self.OnAbout, id = self.menuItemHomeAbout.GetId() )
                self.Bind( wx.EVT_MENU, self.OnHelp, id = self.menuItemHomeHelp.GetId() )
                self.Bind( wx.EVT_MENU, self.OnSetFilename, id = self.menuItemHomeSetFilename.GetId() )
                self.Bind( wx.EVT_MENU, self.OnUpdateLog, id = self.menuItemHomeUpdateLog.GetId() )
                self.Bind( wx.EVT_MENU, self.OnExit, id = self.menuItemHomeExit.GetId() )
                self.Bind( wx.EVT_MENU, self.OnEarlyLeave, id = self.menuItemEarlyEnterData.GetId() )
                self.Bind( wx.EVT_MENU, self.OnEarlyLeaveQueryID, id = self.menuItemEarlyQueryID.GetId() )
                self.Bind( wx.EVT_MENU, self.OnEarlyLeaveQueryDate, id = self.menuItemEarlyQueryDate.GetId() )
                self.Bind( wx.EVT_MENU, self.OnEarlyLeaveQueryName, id = self.menuItemEarlyQueryName.GetId() )
                self.Bind( wx.EVT_MENU, self.OnLatecomers, id = self.menuItemLateEnterData.GetId() )
                self.Bind( wx.EVT_MENU, self.OnLatecomersQueryID, id = self.menuItemLateQueryID.GetId() )
                self.Bind( wx.EVT_MENU, self.OnLatecomersQueryDate, id = self.menuItemLateQueryDate.GetId() )
                self.Bind( wx.EVT_MENU, self.OnLatecomersQueryName, id = self.menuItemLateQueryName.GetId() )
        
        def __del__( self ):
                pass

        #Function to find first empty row in each Excel sheet
        def FindEmpty( self ):
                global earlyleave_empty, latecomers_empty

                #Find last empty row for early leave
                wb = openpyxl.load_workbook(filename)
                sheet = wb["Early Leave"]

                #Iterate through rows in reverse order
                for row in range(sheet.max_row, 0, -1):

                        #When encountered with a row that is filled
                        if sheet['A' + str(row)].value is not None:

                                #Empty row is the row after the last filled row
                                earlyleave_empty = row + 1
                                break

                #Find last empty row for early leave
                wb = openpyxl.load_workbook(filename)
                sheet = wb["Late"]

                #Iterate through rows in reverse order
                for row in range(sheet.max_row, 0, -1):

                        #When encountered with a row that is filled
                        if sheet['A' + str(row)].value is not None:

                                #Empty row is the row after the last filled row
                                latecomers_empty = row + 1
                                break
        
        ## Functions to carry out main functionality
                
        #Function to write email contents
        def WriteEmail(self, filename, student_id, now, reason):        #Now is the variable containing the current datetime
                with open(filename, encoding='utf-8-sig') as csv_file:
                        #Convert csv file into dictionary
                        csv_reader = csv.DictReader(csv_file)

                        selected_row = []

                        for row in csv_reader:
                                if row["Student ID"] == student_id:
                                        selected_row = row

                        #If student not found, return None
                        if selected_row == []:
                                return "Error"
                        
                        #Write Email based on data gathered from CSV file
                        #teachers_emails = f"{selected_row['Email of Form Teacher 1']}, {selected_row['Email of Form Teacher 2']}"
                        teachers_emails = [selected_row['Email of Form Teacher 1'], selected_row['Email of Form Teacher 2']]
                        name_student = selected_row["Name"]

                        #Write email according to format specified in email templates. Select correct file
                        if self.mode == "EarlyLeave":
                                TemplateFile = open("Assets/Email Template - Early Leave.txt", "r")
                        elif self.mode == "Latecomers":
                                TemplateFile = open("Assets/Email Template - Latecomers.txt")

                        #Format text and insert in required variables
                        text = TemplateFile.read()
                        text = text.format(formTeacher_one=selected_row['Name of Form Teacher 1'],\
                                           formTeacher_two=selected_row['Name of Form Teacher 2'],\
                                           name = selected_row['Name'],\
                                           student_class = selected_row['Class'],\
                                           date = now.strftime('%d/%m/%Y'),\
                                           time = now.strftime('%H:%M'),\
                                           reason = reason)

                        #Check if email has been sent successfully
                        if self.SendEmail(teachers_emails, text, name_student) == "Success":
                                return "Success"

        #Function to send email
        def SendEmail(self, to_address, text, name):
                #Get email address and password from text file
                #Read Email Details.txt and get email address and password
                EmailFile = open("Assets/Email Details.txt", "r")
                EmailDetails = EmailFile.read()
                EmailDetails = EmailDetails.split("\n")
                EmailFile.close()
                from_address = str(EmailDetails[0])
                password = str(EmailDetails[1])
                smtp_server = str(EmailDetails[2])
            
                #Startup the server
                server = smtplib.SMTP(smtp_server, 587) 
                server.starttls()
                server.login(from_address,password)

                #Set Email Fields
                msg = MIMEMultipart()
                msg['From'] = from_address
                msg['To'] = ", ".join(to_address)
                
                #Var text allows message in email to change depending on whether the student took early leave or was late for school
                if self.mode == "EarlyLeave":
                        var_text = "leaving early from school"
                elif self.mode == "Latecomers":
                        var_text = "was late for school"
                                
                msg['Subject'] = f"Student {name} {var_text}."
                msg.attach(MIMEText(text, 'plain'))

                #Send Email
                try:
                        server.sendmail(from_address,to_address, msg.as_string())
                        return "Success"
                except PermissionError:
                        #Startup the server
                        server = smtplib.SMTP(smtp_server, 587) 
                        server.starttls()
                        server.login(from_address,password)
                        
                        server.sendmail(from_address,to_address, msg.as_string())
                        return "Success"

        # Virtual event handlers, overide them in your derived class
        def OnSubmit( self, event ):
                global filename, earlyleave_empty, latecomers_empty

                #Check if any field was left blank
                if ((self.textfieldStudentID.GetValue == "") or (self.textfieldReason.GetValue() == "")):
                        self.textConfirmation.SetLabel("Both Student ID Field and Reason Field need to be filled. Please try again.")
                        self.textEmailConfirmation.SetLabel("")
                        return
                        
                #1. Get data (datetime, NRIC and reason) and set confirmation text
                nric = self.textfieldStudentID.GetValue()
                curr_datetime = datetime.datetime.now()
                reason = self.textfieldReason.GetValue()

                #Format curr_datetime to something readable
                formatted_datetime = curr_datetime.strftime('%d/%m/%Y %H:%M')

                #Clear text field
                self.textfieldStudentID.Clear()
                self.textfieldReason.Clear()

                #2. Send Email to Form Teachers
                #If student found and email had been sent successfully
                if self.WriteEmail('Assets/Student Database.csv', nric, curr_datetime, reason) == "Success":
                        self.textConfirmation.SetLabel("")
                        self.textEmailConfirmation.SetLabel("Email has been sent successfully.")
                #If student not found
                else:
                        self.textConfirmation.SetLabel("Student not found. Please try again.")
                        self.textEmailConfirmation.SetLabel("")
                        return

                #3. Get student's name and class from database
                with open("Assets/Student Database.csv", encoding='utf-8-sig') as csv_file:
                    #Convert csv file into dictionary
                    csv_reader = csv.DictReader(csv_file)

                    selected_row = []

                    for row in csv_reader:
                            if row["Student ID"] == nric:
                                    selected_row = row
                                    
                student_name = selected_row["Name"]
                student_class = selected_row["Class"]
                #Censor Student ID when displaying on screen
                censored_id = f"*****{nric[-4:]}"
                
                        
                #4. Open Excel file and select correct sheet according to mode
                wb = openpyxl.load_workbook(filename)
                
                #4. Set last empty row
                if (self.mode == "EarlyLeave"):
                        sheet = wb["Early Leave"]
                        empty = earlyleave_empty
                        earlyleave_empty += 1
                elif (self.mode == "Latecomers"):
                        sheet = wb["Late"]
                        empty = latecomers_empty
                        latecomers_empty += 1

                #5. Add data to Excel file and save
                try:
                    #Set latest row of column A to datetime
                    sheet['A' + str(empty)].value = formatted_datetime
                    #Set latest row of column B to student ID
                    sheet['B' + str(empty)].value = nric
                    #Set latest row of column C to name
                    sheet['C' + str(empty)].value = student_name
                    #Set latest row of column D to class
                    sheet['D' + str(empty)].value = student_class
                    #Set latest row of column E to reason
                    sheet['E' + str(empty)].value = reason
                    wb.save(filename)

                    #Set confirmation text
                    self.textConfirmation.SetLabel(f"The data of student {student_name}, class {student_class} (ID: {censored_id}) \nhas been entered successfully into the system at {formatted_datetime}")
                except:
                    if (self.mode == "EarlyLeave"):
                            earlyleave_empty -= 1
                    elif (self.mode == "Latecomers"):
                            latecomers_empty -= 1
                    #Set confirmation text
                    self.textConfirmation.SetLabel("An error has occurred. Student's data could not be entered into the system.\nPlease close the Data Log Excel File if it's currently open and try again.")
                            

        def OnQueryID( self, event ):
                global filename, earlyleave_empty, latecomers_empty
                if self.textfieldStudentID.GetValue() == "":
                        self.textConfirmation.SetLabel("Student ID Field cannot be left blank. Please try again.")
                        self.textEmailConfirmation.SetLabel("")
                else:   #Run code       
                        
                        #1. Get data (NRIC) and set confirmation text
                        nric = self.textfieldStudentID.GetValue()

                        #Clear text field
                        self.textfieldStudentID.Clear()

                        #2. Get student's name and class from database
                        with open("Assets/Student Database.csv", encoding='utf-8-sig') as csv_file:
                            #Convert csv file into dictionary
                            csv_reader = csv.DictReader(csv_file)

                            selected_row = []

                            for row in csv_reader:
                                    if row["Student ID"] == nric:
                                            selected_row = row
                                            
                        student_name = selected_row["Name"]
                        student_class = selected_row["Class"]

                        #3. Open Excel file and select correct sheet according to mode
                        wb = openpyxl.load_workbook(filename)

                        #4. Set last empty row
                        if (self.mode == "EarlyLeave"):
                                sheet = wb["Early Leave"]
                                empty = earlyleave_empty
                        elif (self.mode == "Latecomers"):
                                sheet = wb["Late"]
                                empty = latecomers_empty

                        #5. Find all data entries corresponding to that student ID
                        entries = []
                        #Iterate through all the rows, sheet is 1-indexed
                        for row in range(1, empty):
                                #If that row contains data entry corresponding to student ID
                                if sheet['B'+str(row)].value == nric:
                                        #Append data entry to a list
                                        entries.append(sheet['A'+str(row)].value)

                        #6. Print out list
                        #If list is empty, then no data entry recorded
                        if len(entries) == 0:
                                self.textConfirmation.SetLabel("Student not found. Please try again.")
                                self.textEmailConfirmation.SetLabel("")
                        #Else print out all the entries
                        else:
                                if self.mode == "EarlyLeave":
                                        text = f"Student {student_name}, class {student_class} (ID: {nric}) \nhas taken early leave from school on these dates:\n"
                                elif self.mode == "Latecomers":
                                        text = f"Student {student_name}, class {student_class} (ID: {nric}) \nhas been late for school on these dates:\n"
                                for item in entries:
                                        text += f"- {str(item)} \n"
                                self.textConfirmation.SetLabel("Data has been retrieved successfully.")

                                #Show Query Dialog
                                QueryDialog(None, text).Show()

        #Function to check legitimacy of date input
        def CheckDate( self, date ):
                try:
                        #Insert date into datetime format
                        datetime.datetime.strptime(date,"%d/%m/%Y")
                        return True
                except:
                        #Incorrect format
                        return False
        
        def OnQueryDate( self, event ):
                global filename, earlyleave_empty, latecomers_empty
                if self.textfieldStudentID.GetValue() == "":
                        self.textConfirmation.SetLabel("Date Field cannot be left blank. Please try again.")
                        self.textEmailConfirmation.SetLabel("")
                else:   #Run code       
                        
                        #1. Get data (requested date) and set confirmation text
                        requested_date = self.textfieldStudentID.GetValue()
                        

                        #Clear text field
                        self.textfieldStudentID.Clear()
                        
                        #Check if that date is really in the specified format
                        if self.CheckDate(requested_date) == False:
                                self.textConfirmation.SetLabel("Date is not in specified format. Please try again.")
                                self.textEmailConfirmation.SetLabel("")
                                return
                        else:
                            #Format date into required format for comparison later (dd/mm/yyyy)
                            #This is to solve issues in comparison such as 01/11/2018 and 1/11/2018
                            requested_date = datetime.datetime.strptime(requested_date, "%d/%m/%Y").strftime("%d/%m/%Y")

                        #2. Open Excel file and select correct sheet according to mode
                        wb = openpyxl.load_workbook(filename)

                        #3. Set last empty row
                        if (self.mode == "EarlyLeave"):
                                sheet = wb["Early Leave"]
                                empty = earlyleave_empty
                        elif (self.mode == "Latecomers"):
                                sheet = wb["Late"]
                                empty = latecomers_empty

                        #4. Find all data entries corresponding to that date
                        entries = []
                        counter = 0
                        #Iterate through all the rows, sheet is 1-indexed, 1st row contains headers, so start on 2nd row
                        for row in range(2, empty):
                            
                                #Retrieve and format date into required format for comparison (dd/mm/yyyy)
                                curr_date = sheet['A'+str(row)].value
                                curr_date = datetime.datetime.strptime(curr_date, "%d/%m/%Y %H:%M").strftime("%d/%m/%Y")

                                #If that row contains data entry corresponding to requested date
                                if curr_date == requested_date:
                                        #Append data entry to a list
                                        entries.append([])
                                        entries[counter].append(sheet['B'+str(row)].value)  #ID
                                        entries[counter].append(sheet['C'+str(row)].value)  #Name  
                                        entries[counter].append(sheet['D'+str(row)].value)  #Class

                        #5. Print out list
                        #If list is empty, then no data entry recorded
                        if len(entries) == 0:
                                self.textConfirmation.SetLabel("No entries for specified date. Please try again.")
                                self.textEmailConfirmation.SetLabel("")
                        #Else print out all the entries
                        else:
                                if self.mode == "EarlyLeave":
                                        text = f"Students who took early leave on {requested_date}:\n"
                                elif self.mode == "Latecomers":
                                        text = f"Students who were late on {requested_date}:\n"
                                for item in entries:
                                        text += f"- {str(item[1])}, class {str(item[2])} (ID: {str(item[0])})\n"
                                self.textConfirmation.SetLabel("Data has been retrieved successfully.")

                                #Show Query Dialog
                                QueryDialog(None, text).Show()

        def OnQueryName( self, event ):
                global filename, earlyleave_empty, latecomers_empty
                if self.textfieldStudentID.GetValue() == "":
                        self.textConfirmation.SetLabel("Name Field cannot be left blank. Please try again.")
                        self.textEmailConfirmation.SetLabel("")
                else:   #Run code       
                        
                        #1. Get data (requested name)
                        requested_name = self.textfieldStudentID.GetValue()
                        
                        #Clear text field
                        self.textfieldStudentID.Clear()

                        #2. Open Excel file and select correct sheet according to mode
                        wb = openpyxl.load_workbook(filename)

                        #3. Set last empty row
                        if (self.mode == "EarlyLeave"):
                                sheet = wb["Early Leave"]
                                empty = earlyleave_empty
                        elif (self.mode == "Latecomers"):
                                sheet = wb["Late"]
                                empty = latecomers_empty

                        #4. Find all data entries corresponding to that name

                        #Format of entries: [[Name, Class, 1st data entry, 2nd data entry, ...], [Name, Class, 1st data entry, 2nd data entry, ...]]
                        entries = []
                        #Iterate through all the rows, sheet is 1-indexed, 1st row contains headers, so start on 2nd row
                        for row in range(2, empty):

                                #Retrieve name of student from current row
                                curr_name = sheet['C'+str(row)].value

                                #If that row contains data entry corresponding to requested name, then add required data
                                if requested_name in curr_name:

                                        #Check if current name is already inputted in
                                        for curr_row in entries:

                                                #curr_row[0] refers to name stored in row that is currently being accessed
                                                #If there exists a list within list of lists that already contains data about current name
                                                if curr_row[0] == curr_name:

                                                        #Add data entry to current row
                                                        curr_row.append(sheet['A'+str(row)].value)  #Date & Time
                                                        break
                                        else:   #If for loop ran without breaking, means requested name is not currently found in entries
                                                #Then, create a new list within list of lists
                                                #Append data entry to a list
                                                entries.append([sheet['C'+str(row)].value, sheet['D'+str(row)].value, sheet['A'+str(row)].value ])  #Name, Class, Date & TIme
                                                
                                                continue    #Continue with next cycle of iteration through rows of Excel file

                                        
                                        
                        #Sort list of entries (By class)
                        entries.sort(key=lambda x: x[1])    #Sort by second element, class

                        
                        #5. Print out list
                        #If list is empty, then no data entry recorded
                        if len(entries) == 0:
                                self.textConfirmation.SetLabel("No entries for specified student. Please try again.")
                                self.textEmailConfirmation.SetLabel("")
                        #Else print out all the entries
                        else:
                                if self.mode == "EarlyLeave":
                                        text = f"Students who contain the name '{requested_name}' and took early leave:\n"
                                elif self.mode == "Latecomers":
                                        text = f"Students who contain the name '{requested_name}' and were late for school:\n"
                                for row in entries:
                                        text += f"\n{row[0]}, class {row[1]}:\n"  #Add header for each unique student
                                        for item in range(2, len(row)):
                                                text += f"- {str(row[item])}\n" #Add individual date & time data entries under each student's header
                                self.textConfirmation.SetLabel("Data has been retrieved successfully.")

                                #Show Query Dialog
                                QueryDialog(None, text).Show()

        ## Functions to show dialogs
        
        def OnAbout( self, event ):
                #Show About Dialog when About menu item selected
                AboutDialog(None).Show()
        
        def OnHelp( self, event ):
                #Show Help Dialog when Help menu item selected
                HelpDialog(None).Show()

        def OnSetFilename( self, event ):
                #Show SetFilename Dialog when SetFilename menu item selected
                SetFilenameDialog(None).Show()
                
        def OnUpdateLog( self, event ):
                #Show UpdateLog Dialog when Update Log menu item selected
                UpdateLogDialog(None).Show()
        
        def OnExit( self, event ):
                #Exit the program
                sys.exit()

        ## Functions to change displays
        
        def OnEarlyLeave( self, event ):
                #Set mode as "EarlyLeave"
                self.mode = "EarlyLeave"

                #Change title to "Early Leave"
                self.textTitle.SetLabel("Early Leave")

                #Show Reason Text and Field
                self.textReason.Show()
                self.textfieldReason.Show()

                #Change instructions text
                self.textEnterStudentID.SetLabel("Enter Student ID Below:")
                self.textReason.SetLabel("Reason for taking Early Leave:")

                #Fix layout
                self.textConfirmation.SetLabel("")
                self.textEmailConfirmation.SetLabel("")
                self.Layout()

                #Rebind submit button to Submit function
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnSubmit )

        def OnEarlyLeaveQueryID( self, event ):
                #Set mode as "EarlyLeave"
                self.mode = "EarlyLeave"

                #Change title to "Early Leave: Query by Student ID"
                self.textTitle.SetLabel("Early Leave:\nQuery by Student ID")

                #Hide Reason Text and Field
                self.textReason.Hide()
                self.textfieldReason.Hide()

                #Change instructions text
                self.textEnterStudentID.SetLabel("Enter Student ID Below:")

                #Fix layout
                self.textConfirmation.SetLabel("")
                self.textEmailConfirmation.SetLabel("")
                self.Layout()

                #Rebind submit button to query ID function
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnQueryID )
        
        def OnEarlyLeaveQueryDate( self, event ):
                #Set mode as "EarlyLeave"
                self.mode = "EarlyLeave"

                #Change title to "Early Leave: Query by Date"
                self.textTitle.SetLabel("Early Leave:\nQuery by Date")

                #Hide Reason Text and Field
                self.textReason.Hide()
                self.textfieldReason.Hide()

                #Change instructions text
                self.textEnterStudentID.SetLabel("Enter date in the format dd/mm/yyyy:")
                
                #Fix layout
                self.textConfirmation.SetLabel("")
                self.textEmailConfirmation.SetLabel("")
                self.Layout()

                #Rebind submit button to query date function
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnQueryDate )

        def OnEarlyLeaveQueryName( self, event ):
                #Set mode as "EarlyLeave"
                self.mode = "EarlyLeave"

                #Change title to "Early Leave: Query by Name"
                self.textTitle.SetLabel("Early Leave:\nQuery by Name")

                #Hide Reason Text and Field
                self.textReason.Hide()
                self.textfieldReason.Hide()

                #Change instructions text
                self.textEnterStudentID.SetLabel("Enter Name of Student Below:")
                
                #Fix layout
                self.textConfirmation.SetLabel("")
                self.textEmailConfirmation.SetLabel("")
                self.Layout()

                #Rebind submit button to query date function
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnQueryName )
        
        def OnLatecomers( self, event ):
                #Set mode as "Latecomers" 
                self.mode = "Latecomers"

                #Change title to "Latecomers"
                self.textTitle.SetLabel("Latecomers")

                #Show Reason Text and Field
                self.textReason.Show()
                self.textfieldReason.Show()

                #Change instructions text
                self.textEnterStudentID.SetLabel("Enter Student ID Below:")
                self.textReason.SetLabel("Reason for Latecoming:")
                
                #Fix layout
                self.textConfirmation.SetLabel("")
                self.textEmailConfirmation.SetLabel("")
                self.Layout()

                #Rebind submit button to Submit function
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnSubmit )

        def OnLatecomersQueryID( self, event ):
                #Set mode as "Latecomers"
                self.mode = "Latecomers"

                #Change title to "Latecomers: Query by Student ID"
                self.textTitle.SetLabel("Latecomers:\nQuery by Student ID")

                #Hide Reason Text and Field
                self.textReason.Hide()
                self.textfieldReason.Hide()

                #Change instructions text
                self.textEnterStudentID.SetLabel("Enter Student ID Below:")

                #Fix layout
                self.textConfirmation.SetLabel("")
                self.textEmailConfirmation.SetLabel("")
                self.Layout()

                #Rebind submit button to query ID function
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnQueryID )

        def OnLatecomersQueryDate( self, event ):
                #Set mode as "Latecomers"
                self.mode = "Latecomers"

                #Change title to "Latecomers: Query by Date"
                self.textTitle.SetLabel("Latecomers:\nQuery by Date")

                #Hide Reason Text and Field
                self.textReason.Hide()
                self.textfieldReason.Hide()

                #Change instructions text
                self.textEnterStudentID.SetLabel("Enter date in the format dd/mm/yyyy:")

                #Fix layout
                self.textConfirmation.SetLabel("")
                self.textEmailConfirmation.SetLabel("")
                self.Layout()

                #Rebind submit button to query date function
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnQueryDate )

        def OnLatecomersQueryName( self, event ):
                #Set mode as "Latecomers"
                self.mode = "Latecomers"

                #Change title to "Latecomers: Query by Name"
                self.textTitle.SetLabel("Latecomers:\nQuery by Name")

                #Hide Reason Text and Field
                self.textReason.Hide()
                self.textfieldReason.Hide()

                #Change instructions text
                self.textEnterStudentID.SetLabel("Enter Name of Student Below:")

                #Fix layout
                self.textConfirmation.SetLabel("")
                self.textEmailConfirmation.SetLabel("")
                self.Layout()

                #Rebind submit button to query date function
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnQueryName )
        

###########################################################################
## Class AboutDialog
###########################################################################

class AboutDialog ( wx.Dialog ):
        
        def __init__( self, parent ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"About", pos = wx.DefaultPosition, size = wx.Size( 400,400 ), style = wx.DEFAULT_DIALOG_STYLE )
                
                self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )
                
                bSizerAbout = wx.BoxSizer( wx.VERTICAL )

                #Read About - Version.txt and add it to the dialog
                AboutVersion = open("Assets/About - Version.txt", "r")
                text = AboutVersion.read()
                AboutVersion.close()
                
                self.textAbout = wx.StaticText( self, wx.ID_ANY, text, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textAbout.Wrap( -1 )
                self.textAbout.SetFont( wx.Font( 13, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerAbout.Add( self.textAbout, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 20 )

                #Read About - Features.txt and add it to the dialog
                AboutFeatures = open("Assets/About - Features.txt", "r")
                text = AboutFeatures.read()
                AboutFeatures.close()
                
                self.textFeatures = wx.StaticText( self, wx.ID_ANY, text, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textFeatures.Wrap( -1 )
                self.textFeatures.SetFont( wx.Font( 13, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerAbout.Add( self.textFeatures, 0, wx.ALL, 10 )
                
                
                self.SetSizer( bSizerAbout )
                self.Layout()
                
                #Added self.Fit()
                self.Fit()
                
                self.Centre( wx.BOTH )
        
        def __del__( self ):
                pass
        

###########################################################################
## Class HelpDialog
###########################################################################

class HelpDialog ( wx.Dialog ):
        
        def __init__( self, parent ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"Help", pos = wx.DefaultPosition, size = wx.Size( 500,300 ), style = wx.DEFAULT_DIALOG_STYLE )
                
                self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )
                
                bSizerHelp = wx.BoxSizer( wx.VERTICAL )
                
                self.textHelpTitle = wx.StaticText( self, wx.ID_ANY, u"Help", wx.DefaultPosition, wx.Size( -1,-1 ), 0 )
                self.textHelpTitle.Wrap( -1 )
                self.textHelpTitle.SetFont( wx.Font( 25, 70, 90, 92, False, wx.EmptyString ) )
                
                bSizerHelp.Add( self.textHelpTitle, 0, wx.ALIGN_LEFT|wx.ALL, 15 )

                #Read Help.txt and add it to the dialog
                HelpFile = open("Assets/Help.txt", "r")
                text = HelpFile.read()
                HelpFile.close()
                
                self.textHelpInstructions = wx.StaticText( self, wx.ID_ANY, text, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textHelpInstructions.Wrap( -1 )
                self.textHelpInstructions.SetFont( wx.Font( 13, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerHelp.Add( self.textHelpInstructions, 0, wx.ALIGN_LEFT|wx.ALL, 10 )
                
                self.SetSizer( bSizerHelp )
                self.Layout()
                
                #Added self.Fit()
                self.Fit()
                
                self.Centre( wx.BOTH )
        
        def __del__( self ):
                pass
        

###########################################################################
## Class UpdateLogDialog
###########################################################################

class UpdateLogDialog ( wx.Dialog ):
        
        def __init__( self, parent ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"Help", pos = wx.DefaultPosition, size = wx.Size( 750,500 ), style = wx.DEFAULT_DIALOG_STYLE )
                
                self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

                #Read Update Log.txt and add it to the dialog
                UpdateLog = open("Assets/Update Log.txt", "r")
                text = UpdateLog.read()
                UpdateLog.close()
                
                #Setup panel
                self.panel = ScrollPanel(self, "Update Log", text)

                #Put panel into box sizer and display the panel
                bSizerUpdateLog = wx.BoxSizer( wx.VERTICAL )
                bSizerUpdateLog.Add(self.panel, 1, wx.EXPAND, 10)

                self.SetSizer(bSizerUpdateLog)
                self.Layout()
                
        def __del__( self ):
                pass

###########################################################################
## Class SetFilenameDialog
###########################################################################
        
class SetFilenameDialog ( wx.Dialog ):

        def __init__( self, parent ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"Set Filename", pos = wx.DefaultPosition, size = wx.DefaultSize, style = wx.DEFAULT_DIALOG_STYLE )
                
                self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )
                
                bSizerSetFilename = wx.BoxSizer( wx.VERTICAL )
                
                self.textSetFilenameTitle = wx.StaticText( self, wx.ID_ANY, u"Set Filename", wx.DefaultPosition, wx.Size( -1,-1 ), 0 )
                self.textSetFilenameTitle.Wrap( -1 )
                self.textSetFilenameTitle.SetFont( wx.Font( 15, 70, 90, 92, False, wx.EmptyString ) )
                
                bSizerSetFilename.Add( self.textSetFilenameTitle, 0, wx.ALIGN_LEFT|wx.ALL, 15 )
                
                self.textSetFilenameInstructions = wx.StaticText( self, wx.ID_ANY, u"Type filename of Excel File to store student data in (remember to include '.xlsx'):\n(Ensure that the file is stored in the 'Assets' Folder)", wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textSetFilenameInstructions.Wrap( -1 )
                self.textSetFilenameInstructions.SetFont( wx.Font( 13, 70, 90, wx.FONTWEIGHT_NORMAL, False, wx.EmptyString ) )
                bSizerSetFilename.Add( self.textSetFilenameInstructions, 0, wx.ALIGN_LEFT|wx.ALL, 10 )

                self.textfieldSetFilename = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
                bSizerSetFilename.Add( self.textfieldSetFilename, 0, wx.ALIGN_TOP|wx.ALL|wx.EXPAND, 10 )

                self.buttonSubmit = wx.Button( self, wx.ID_ANY, u"Submit", wx.DefaultPosition, wx.DefaultSize, 0 )
                bSizerSetFilename.Add( self.buttonSubmit, 0, wx.ALL|wx.EXPAND, 10 )

                self.textConfirmation = wx.StaticText( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
                self.textConfirmation.Wrap( -1 )
                bSizerSetFilename.Add( self.textConfirmation, 0, wx.ALIGN_CENTER|wx.ALIGN_TOP|wx.ALL|wx.EXPAND, 10 )
                
                self.SetSizer( bSizerSetFilename )
                self.Layout()
                
                #Added self.Fit()
                self.Fit()
                
                self.Centre( wx.BOTH )

                # Connect Events
                self.buttonSubmit.Bind( wx.EVT_BUTTON, self.OnSubmitFilename )
                
        def OnSubmitFilename( self, event ):
                global filename

                #If text field is empty, output an error message
                if self.textfieldSetFilename.GetValue() == "":
                        self.textConfirmation.SetLabel("Filename field cannot be left blank. Please try again.")
                else:
                        #Run Code
                        # Get filename from text field
                        filename = f"Assets/{self.textfieldSetFilename.GetValue()}"
                        file = open("Assets/Filename.txt", "w")
                        file.write(filename)

                        # Clear text field
                        self.textfieldSetFilename.Clear()

                        # Display confirmation text
                        self.textConfirmation.SetLabel("Filename '" + filename + "' has been set.")

###########################################################################
## Class UpdateLogDialog
###########################################################################

class QueryDialog ( wx.Dialog ):
        
        def __init__( self, parent, text ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"Query Info", pos = wx.DefaultPosition, size = wx.Size( 700,500 ), style = wx.DEFAULT_DIALOG_STYLE )
                
                self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )

                #Setup panel
                self.panel = ScrollPanel(self, "Query Info", text)

                #Put panel into box sizer and display the panel
                bSizerQueryDialog = wx.BoxSizer( wx.VERTICAL )
                bSizerQueryDialog.Add(self.panel, 1, wx.EXPAND, 10)

                self.SetSizer(bSizerQueryDialog)
                self.Layout()
        
        def __del__( self ):
                pass
                        
#Run Code (allows app to be run)
class MainApp(wx.App):
        def OnInit(self):
                mainFrame = MainFrame(None)
                mainFrame.Show(True)
                return True

if __name__ == '__main__':
        app = MainApp()
        #Loop running of app, if not it will instantly disappear
        app.MainLoop()
